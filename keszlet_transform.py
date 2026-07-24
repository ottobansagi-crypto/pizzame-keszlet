#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
KÉSZLET FELTÖLTÉS – nyers Wildom export -> V31-kompatibilis 'Készlet feltöltés' lap.

Kétféleképpen futtatható:

  1) Mappa-mód (GitHub Actionshöz, ez az alapértelmezett):
         python keszlet_transform.py
     A 'bemenet/' mappában lévő két .xlsx-et automatikusan felismeri
     (HU vs SK az oszlopszám alapján), és a 'kimenet/' mappába ír egy
     keszlet_feltoltes_ÉÉÉÉ-HH-NN.xlsx fájlt.

  2) Kézi mód (explicit fájlok):
         python keszlet_transform.py HU.xlsx SK.xlsx OUTPUT.xlsx

Kulcs elv:
  A nyers export 'Sort Index' oszlopa == V31 termékindex (1..597). Ezen a kulcson
  történik a párosítás – nincs szükség kézi 'Előfordulás' szűrésre / rejtésre.

Üzletoszlop-térkép (config.json):
  - HU: az export B..(Debrecen) oszlopai pozíció szerint a V31 kódokra (hu_codes).
  - SK: Michalska->E03 (BR1), Kolárska->E04 (BR2).
  - SK italok: a Wildom a KÖZÖS indexen adja őket (pl. Coca-Cola=162), de a V31-ben
    a BR oszlopoknál a '(sk)' duplikátum sorba kell kerülniük -> sk_remap.
  - Mom (E09) és Terpolit (E10): nincs nyers forrás -> üres.
"""
import sys, json, os, glob, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
CONFIG = json.load(open(os.path.join(HERE, 'config.json'), encoding='utf-8'))

MASTER    = {int(k): v for k, v in CONFIG['master'].items()}
HU_CODES  = CONFIG['hu_codes']
SK_REMAP  = {int(k): int(v) for k, v in CONFIG['sk_remap'].items()}
SK_STORES = CONFIG['sk_stores']
LAYOUT    = [(l, c, d) for (l, c, d) in CONFIG['final_layout'] if c]

report = {'hu_blank_idx': [], 'sk_blank_idx': [], 'negatives': [],
          'sk_dupidx_in_hu': [], 'unknown_idx': [], 'notes': []}


def read_raw(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    return rows[0], rows[1:]


def num(v):
    if v is None or (isinstance(v, str) and v.strip() == ''):
        return None
    return v


def classify(path):
    """HU vs SK felismerése oszlopszám alapján (SK ~5 oszlop, HU ~47)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ncol = wb.active.max_column
    wb.close()
    return 'SK' if ncol <= 12 else 'HU'


def find_inputs():
    files = sorted(glob.glob(os.path.join(HERE, 'bemenet', '*.xlsx')))
    files = [f for f in files if not os.path.basename(f).startswith('~$')]
    hu = sk = None
    for f in files:
        kind = classify(f)
        if kind == 'HU' and hu is None:
            hu = f
        elif kind == 'SK' and sk is None:
            sk = f
    return hu, sk


def build(hu_path, sk_path):
    grid = {}

    # ---------- HU ----------
    hu_header, hu_rows = read_raw(hu_path)
    n_store = len(hu_header) - 3
    if n_store != len(HU_CODES):
        report['notes'].append(
            f"FIGYELEM: a HU export {n_store} üzletoszlopot tartalmaz, a térkép {len(HU_CODES)}-t vár. "
            f"Ellenőrizd az üzletkijelölést (Debrecen legyen az utolsó).")
    si_col = len(hu_header) - 1
    for row in hu_rows:
        if row[0] is None:
            continue
        si = num(row[si_col])
        if not isinstance(si, (int, float)):
            report['hu_blank_idx'].append(row[0]); continue
        si = int(si)
        if si not in MASTER:
            report['unknown_idx'].append(('HU', si, row[0])); continue
        for j, code in enumerate(HU_CODES):
            v = num(row[1 + j])
            if v is None:
                continue
            grid[(si, code)] = v
            if isinstance(v, (int, float)) and v < 0:
                report['negatives'].append((si, code, v, MASTER[si]['name']))
        if si in SK_REMAP.values():
            report['sk_dupidx_in_hu'].append((si, row[0]))

    # ---------- SK ----------
    sk_header, sk_rows = read_raw(sk_path)
    sk_si_col = len(sk_header) - 1
    sk_col_code = {i: SK_STORES[h] for i, h in enumerate(sk_header) if h in SK_STORES}
    if not sk_col_code:
        report['notes'].append("FIGYELEM: az SK üzletoszlopok nem azonosíthatók a fejlécből.")
    for row in sk_rows:
        if row[0] is None:
            continue
        si = num(row[sk_si_col])
        if not isinstance(si, (int, float)):
            report['sk_blank_idx'].append(row[0]); continue
        si = int(si)
        if si not in MASTER:
            report['unknown_idx'].append(('SK', si, row[0])); continue
        target = SK_REMAP.get(si, si)
        for cidx, code in sk_col_code.items():
            v = num(row[cidx])
            if v is None:
                continue
            grid[(target, code)] = v
            if isinstance(v, (int, float)) and v < 0:
                report['negatives'].append((target, code, v, MASTER[target]['name']))
    return grid


def write(grid, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Készlet feltöltés'
    A = 'Arial'
    hdr = Font(name=A, bold=True, size=10)
    base = Font(name=A, size=10)
    mod_fill = {'A': 'DCE6F1', 'B': 'EBF1DE', 'C': 'FDE9D9', 'D': 'E4DFEC', 'E': 'F2DCDB'}

    ws['A1'] = 'Termékek'; ws['A1'].font = hdr
    ws['B1'] = 'Negatív értékek ennyi oszlopban'; ws['B1'].font = hdr
    ws['B1'].alignment = Alignment(wrap_text=True, horizontal='center')
    for col, code, dname in LAYOUT:
        c1 = ws[f'{col}1']; c1.value = code; c1.font = hdr
        c1.alignment = Alignment(horizontal='center')
        if code[0] in mod_fill:
            c1.fill = PatternFill('solid', fgColor=mod_fill[code[0]])
        c2 = ws[f'{col}2']; c2.value = dname; c2.font = base
        c2.alignment = Alignment(horizontal='center')
    max_idx = max(MASTER)
    ws['B3'] = f'=SUM(B4:B{max_idx + 3})'; ws['B3'].font = hdr

    last_col = LAYOUT[-1][0]
    for idx in range(1, max_idx + 1):
        r = idx + 3
        m = MASTER.get(idx)
        ws.cell(r, 1, (m and m['name']) or f'(idx {idx})').font = base
        ws.cell(r, 2, f'=COUNTIF(C{r}:{last_col}{r},"<0")').font = base
        for col, code, _ in LAYOUT:
            v = grid.get((idx, code))
            if v is not None:
                ws[f'{col}{r}'].value = v
                ws[f'{col}{r}'].font = base

    ws.freeze_panes = 'C4'
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 12

    lg = wb.create_sheet('Napló')
    r = [1]
    def put(txt, bold=False, size=10):
        lg.cell(r[0], 1, txt).font = Font(name=A, bold=bold, size=size); r[0] += 1
    put('ELLENŐRZÉSI NAPLÓ – automatikus készlet feltöltés', bold=True, size=12)
    put(f'Generálva: {datetime.datetime.now():%Y-%m-%d %H:%M}'); r[0] += 1
    if report['unknown_idx']:
        put('!!! ISMERETLEN INDEX (nincs a V31 masterben – config frissítés kellhet):', bold=True)
        for src, idx, nm in report['unknown_idx']:
            put(f'   [{src}] idx {idx} | {nm}')
        r[0] += 1
    put(f"Index nélküli HU sorok (kimaradtak) – {len(report['hu_blank_idx'])} db:", bold=True)
    for x in report['hu_blank_idx']:
        put(f'   {x}')
    r[0] += 1
    put(f"Index nélküli SK sorok (kimaradtak) – {len(report['sk_blank_idx'])} db:", bold=True)
    for x in report['sk_blank_idx']:
        put(f'   {x}')
    r[0] += 1
    put(f"Negatív értékek (hó végén javítandó) – {len(report['negatives'])} db:", bold=True)
    for idx, code, v, nm in report['negatives']:
        put(f'   idx {idx} | {code} | {v} | {nm}')
    if report['sk_dupidx_in_hu']:
        r[0] += 1
        put('FIGYELEM: HU adat (sk)-duplikátum indexen:', bold=True)
        for idx, nm in report['sk_dupidx_in_hu']:
            put(f'   idx {idx} | {nm}')
    if report['notes']:
        r[0] += 1
        put('Egyéb megjegyzések:', bold=True)
        for n in report['notes']:
            put(f'   {n}')
    lg.column_dimensions['A'].width = 95

    wb.save(out_path)


def main():
    if len(sys.argv) == 4:
        hu_path, sk_path, out_path = sys.argv[1], sys.argv[2], sys.argv[3]
    else:
        hu_path, sk_path = find_inputs()
        if not hu_path or not sk_path:
            sys.exit("HIBA: a 'bemenet/' mappában nem található HU és SK .xlsx is. "
                     "Tölts fel pontosan egy HU és egy SK exportot.")
        out_dir = os.path.join(HERE, 'kimenet')
        os.makedirs(out_dir, exist_ok=True)
        out_path = os.path.join(out_dir, f'keszlet_feltoltes_{datetime.date.today():%Y-%m-%d}.xlsx')

    print(f"HU bemenet: {os.path.basename(hu_path)}")
    print(f"SK bemenet: {os.path.basename(sk_path)}")
    grid = build(hu_path, sk_path)
    write(grid, out_path)
    print(f"Kész: {out_path}")
    print(f"  kitöltött cella: {len(grid)}")
    print(f"  HU index nélküli: {len(report['hu_blank_idx'])} | SK index nélküli: {len(report['sk_blank_idx'])}")
    print(f"  negatív cella: {len(report['negatives'])}")
    if report['unknown_idx']:
        print(f"  !!! ISMERETLEN INDEX: {len(report['unknown_idx'])} – nézd a Napló lapot (config frissítés kellhet)")


if __name__ == '__main__':
    main()
